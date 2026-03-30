"""
Session Aggregator — GUI
Merges Keyboard Activity Tracker JSON sessions into a formatted Excel report.
Matches the dark navy theme of the main tracker application.
"""

import sys
import os
import json
import threading
from pathlib import Path
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFileDialog, QProgressBar,
    QTextEdit, QFrame, QSizePolicy, QScrollArea, QGridLayout
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QTimer
from PyQt5.QtGui import (
    QFont, QColor, QPalette, QIcon, QPixmap, QPainter, QBrush,
    QPen, QLinearGradient
)

# ── Colour palette (matches main tracker) ─────────────────────────────────────
BG_DARK       = "#0F1318"
BG_CARD       = "#1A1F2E"
BG_INPUT      = "#252B3B"
BG_HOVER      = "#2D3447"
ACCENT        = "#4C9BE8"
ACCENT_DARK   = "#3A7BC8"
ACCENT_GREEN  = "#4CAF82"
ACCENT_RED    = "#E84C4C"
ACCENT_ORANGE = "#E8A84C"
TEXT_PRIMARY  = "#E8EBF0"
TEXT_SECONDARY= "#A0ABBF"
TEXT_MUTED    = "#5A6478"
BORDER        = "#2D3447"
BORDER_ACCENT = "#4C9BE8"

STYLESHEET = f"""
QMainWindow, QWidget {{
    background-color: {BG_DARK};
    color: {TEXT_PRIMARY};
    font-family: 'Segoe UI';
    font-size: 13px;
}}

/* ── Cards ── */
#card {{
    background-color: {BG_CARD};
    border: 1px solid {BORDER};
    border-radius: 10px;
}}

/* ── Section labels ── */
#section_label {{
    color: {TEXT_SECONDARY};
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
}}

/* ── Input fields ── */
QLineEdit {{
    background-color: {BG_INPUT};
    border: 1px solid {BORDER};
    border-radius: 6px;
    color: {TEXT_PRIMARY};
    padding: 8px 12px;
    font-size: 13px;
    selection-background-color: {ACCENT};
}}
QLineEdit:focus {{
    border: 1px solid {ACCENT};
    background-color: {BG_HOVER};
}}
QLineEdit::placeholder {{
    color: {TEXT_MUTED};
}}

/* ── Browse buttons ── */
#browse_btn {{
    background-color: {BG_HOVER};
    border: 1px solid {BORDER};
    border-radius: 6px;
    color: {TEXT_SECONDARY};
    padding: 8px 16px;
    font-size: 12px;
    font-weight: 600;
    min-width: 90px;
}}
#browse_btn:hover {{
    background-color: {ACCENT};
    border-color: {ACCENT};
    color: white;
}}
#browse_btn:pressed {{
    background-color: {ACCENT_DARK};
}}

/* ── Generate button ── */
#generate_btn {{
    background-color: {ACCENT};
    border: none;
    border-radius: 8px;
    color: white;
    padding: 12px 28px;
    font-size: 14px;
    font-weight: 700;
    min-height: 44px;
}}
#generate_btn:hover {{
    background-color: {ACCENT_DARK};
}}
#generate_btn:pressed {{
    background-color: #2D6AB0;
}}
#generate_btn:disabled {{
    background-color: {BG_HOVER};
    color: {TEXT_MUTED};
}}

/* ── Open button ── */
#open_btn {{
    background-color: {BG_HOVER};
    border: 1px solid {ACCENT_GREEN};
    border-radius: 8px;
    color: {ACCENT_GREEN};
    padding: 12px 28px;
    font-size: 14px;
    font-weight: 700;
    min-height: 44px;
}}
#open_btn:hover {{
    background-color: {ACCENT_GREEN};
    color: white;
}}
#open_btn:disabled {{
    background-color: {BG_HOVER};
    border-color: {BORDER};
    color: {TEXT_MUTED};
}}

/* ── Progress bar ── */
QProgressBar {{
    background-color: {BG_INPUT};
    border: 1px solid {BORDER};
    border-radius: 4px;
    height: 8px;
    text-align: center;
    color: transparent;
}}
QProgressBar::chunk {{
    background-color: {ACCENT};
    border-radius: 4px;
}}

/* ── Log area ── */
QTextEdit {{
    background-color: {BG_INPUT};
    border: 1px solid {BORDER};
    border-radius: 6px;
    color: {TEXT_SECONDARY};
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 12px;
    padding: 8px;
}}

/* ── Session preview chips ── */
#stat_card {{
    background-color: {BG_INPUT};
    border: 1px solid {BORDER};
    border-radius: 8px;
    padding: 4px;
}}
#stat_value {{
    color: {ACCENT};
    font-size: 22px;
    font-weight: 700;
}}
#stat_label {{
    color: {TEXT_MUTED};
    font-size: 11px;
}}

/* ── Divider ── */
#divider {{
    background-color: {BORDER};
    max-height: 1px;
    min-height: 1px;
}}

/* ── Scrollbar ── */
QScrollBar:vertical {{
    background: {BG_DARK};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {BG_HOVER};
    border-radius: 4px;
    min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{
    background: {ACCENT};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0px;
}}
"""


# ── Worker thread ─────────────────────────────────────────────────────────────
class ReportWorker(QObject):
    progress  = pyqtSignal(int)
    log       = pyqtSignal(str)
    finished  = pyqtSignal(str)   # output path
    error     = pyqtSignal(str)

    def __init__(self, input_dir, output_path, username):
        super().__init__()
        self.input_dir   = input_dir
        self.output_path = output_path
        self.username    = username

    def run(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # ── helpers ──────────────────────────────────────────────────────
            C_HEADER_BG  = "1A1F2E"
            C_HEADER_FG  = "FFFFFF"
            C_SUBHEAD_BG = "2D3447"
            C_ALT_ROW    = "F4F6FA"
            C_TOTAL_BG   = "E8EDF5"
            C_ACCENT     = "4C9BE8"

            def _side():
                return Side(style="thin", color="D0D7E3")
            def _border():
                s = _side()
                return Border(left=s, right=s, top=s, bottom=s)
            def _hf():
                return Font(name="Arial", bold=True, color=C_HEADER_FG, size=10)
            def _bf(bold=False):
                return Font(name="Arial", bold=bold, size=10)
            def _hfill():
                return PatternFill("solid", start_color=C_HEADER_BG)
            def _sfill():
                return PatternFill("solid", start_color=C_SUBHEAD_BG)
            def _afill():
                return PatternFill("solid", start_color=C_ALT_ROW)
            def _tfill():
                return PatternFill("solid", start_color=C_TOTAL_BG)
            def _ctr():
                return Alignment(horizontal="center", vertical="center", wrap_text=True)

            def apply_header(ws, row, values):
                for i, v in enumerate(values):
                    c = ws.cell(row=row, column=i+1, value=v)
                    c.font = _hf(); c.fill = _hfill()
                    c.alignment = _ctr(); c.border = _border()

            def apply_row(ws, row, values, alt=False, total=False):
                fill = _tfill() if total else (_afill() if alt else None)
                for i, v in enumerate(values):
                    c = ws.cell(row=row, column=i+1, value=v)
                    c.font = _bf(bold=total); c.alignment = _ctr()
                    c.border = _border()
                    if fill: c.fill = fill

            def set_widths(ws, widths):
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            def title_row(ws, cols, text):
                ws.merge_cells(f"A1:{get_column_letter(cols)}1")
                t = ws["A1"]
                t.value = text
                t.font = Font(name="Arial", bold=True, color=C_ACCENT, size=12)
                t.alignment = _ctr()
                t.fill = _sfill()
                ws.row_dimensions[1].height = 28

            def fmt_dt(iso):
                try: return datetime.fromisoformat(iso).strftime("%Y-%m-%d %H:%M:%S")
                except: return iso
            def fmt_date(iso):
                try: return datetime.fromisoformat(iso).strftime("%Y-%m-%d")
                except: return iso

            # ── load sessions ─────────────────────────────────────────────────
            self.log.emit("📂  Scanning input folder...")
            self.progress.emit(5)

            files = sorted(Path(self.input_dir).glob("*.json"))
            if not files:
                self.error.emit("No JSON files found in the selected folder.")
                return

            sessions = []
            for f in files:
                try:
                    data = json.loads(f.read_text(encoding="utf-8"))
                    if data.get("encrypted"):
                        self.log.emit(f"⚠️  Skipping encrypted file: {f.name}")
                        continue
                    sessions.append(data)
                    self.log.emit(f"✅  Loaded: {f.name}")
                except Exception as e:
                    self.log.emit(f"⚠️  Could not read {f.name}: {e}")

            if not sessions:
                self.error.emit("No readable session files found.")
                return

            sessions.sort(key=lambda s: s.get("session_start", ""))
            self.log.emit(f"\n📊  {len(sessions)} session(s) loaded — building workbook...")
            self.progress.emit(15)

            wb = Workbook()
            wb.remove(wb.active)

            # ── Sheet 1: Summary ──────────────────────────────────────────────
            self.log.emit("📝  Writing Summary sheet...")
            ws = wb.create_sheet("Summary")
            ws.sheet_view.showGridLines = False
            title_row(ws, 12, f"Keyboard Activity Tracker — Session Summary   |   User: {self.username}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            hdrs = ["Session Date","Session Start","Session End","Duration","Total Keys","Char Keys","Modifier Keys","Special Keys","Active Time","Idle Time","Efficiency %","Breaks"]
            apply_header(ws, 2, hdrs)
            ws.row_dimensions[2].height = 22
            for i, s in enumerate(sessions):
                apply_row(ws, 3+i, [
                    fmt_date(s.get("session_start","")),
                    fmt_dt(s.get("session_start","")),
                    fmt_dt(s.get("session_end", s.get("snapshot_time",""))),
                    s.get("total_duration_formatted", s.get("session_duration_formatted","")),
                    s.get("keystroke_counts",{}).get("total",0),
                    s.get("keystroke_counts",{}).get("character_keys",0),
                    s.get("keystroke_counts",{}).get("modifier_keys",0),
                    s.get("keystroke_counts",{}).get("special_keys",0),
                    s.get("time_metrics",{}).get("active_formatted",""),
                    s.get("time_metrics",{}).get("idle_formatted",""),
                    s.get("time_metrics",{}).get("typing_efficiency_percent",0),
                    s.get("break_count",0),
                ], alt=(i%2==1))
            n = len(sessions)
            apply_row(ws, 3+n, [
                "", f"TOTALS / AVG ({n} sessions)", "", "",
                sum(s.get("keystroke_counts",{}).get("total",0) for s in sessions),
                sum(s.get("keystroke_counts",{}).get("character_keys",0) for s in sessions),
                sum(s.get("keystroke_counts",{}).get("modifier_keys",0) for s in sessions),
                sum(s.get("keystroke_counts",{}).get("special_keys",0) for s in sessions),
                "", "",
                round(sum(s.get("time_metrics",{}).get("typing_efficiency_percent",0) for s in sessions)/n, 2),
                sum(s.get("break_count",0) for s in sessions),
            ], total=True)
            set_widths(ws, [14,20,20,12,12,12,14,14,12,12,13,8])
            ws.freeze_panes = "A3"
            self.progress.emit(30)

            # ── Sheet 2: Keystroke Detail ──────────────────────────────────────
            self.log.emit("📝  Writing Keystroke Detail sheet...")
            ws = wb.create_sheet("Keystroke Detail")
            ws.sheet_view.showGridLines = False
            title_row(ws, 8, "Keystroke Category Breakdown — Per Session")
            apply_header(ws, 2, ["Session Date","Total Keys","Character Keys","Space Keys","Backspace Keys","Enter Keys","Modifier Keys","Special Keys"])
            ws.row_dimensions[2].height = 22
            for i, s in enumerate(sessions):
                kc = s.get("keystroke_counts",{})
                apply_row(ws, 3+i, [fmt_date(s.get("session_start","")), kc.get("total",0), kc.get("character_keys",0), kc.get("space_keys",0), kc.get("backspace_keys",0), kc.get("enter_keys",0), kc.get("modifier_keys",0), kc.get("special_keys",0)], alt=(i%2==1))
            apply_row(ws, 3+n, ["TOTALS", sum(s.get("keystroke_counts",{}).get("total",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("character_keys",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("space_keys",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("backspace_keys",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("enter_keys",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("modifier_keys",0) for s in sessions), sum(s.get("keystroke_counts",{}).get("special_keys",0) for s in sessions)], total=True)
            set_widths(ws, [14,12,14,12,15,12,14,13])
            ws.freeze_panes = "A3"
            self.progress.emit(48)

            # ── Sheet 3: KPM Metrics ───────────────────────────────────────────
            self.log.emit("📝  Writing KPM Metrics sheet...")
            ws = wb.create_sheet("KPM Metrics")
            ws.sheet_view.showGridLines = False
            title_row(ws, 4, "Keys Per Minute (KPM) Metrics — Per Session")
            apply_header(ws, 2, ["Session Date","Average KPM","Peak KPM","End-of-Session KPM"])
            ws.row_dimensions[2].height = 22
            for i, s in enumerate(sessions):
                km = s.get("kpm_metrics",{})
                apply_row(ws, 3+i, [fmt_date(s.get("session_start","")), km.get("average_kpm",0), km.get("peak_kpm",0), km.get("current_kpm",0)], alt=(i%2==1))
            apply_row(ws, 3+n, ["AVERAGES", round(sum(s.get("kpm_metrics",{}).get("average_kpm",0) for s in sessions)/n,2), round(sum(s.get("kpm_metrics",{}).get("peak_kpm",0) for s in sessions)/n,2), ""], total=True)
            set_widths(ws, [14,15,12,20])
            ws.freeze_panes = "A3"
            self.progress.emit(62)

            # ── Sheet 4: Time Metrics ──────────────────────────────────────────
            self.log.emit("📝  Writing Time Metrics sheet...")
            ws = wb.create_sheet("Time Metrics")
            ws.sheet_view.showGridLines = False
            title_row(ws, 6, "Active / Idle Time — Per Session")
            apply_header(ws, 2, ["Session Date","Session Duration","Active Time","Idle Time","Efficiency %","Break Count"])
            ws.row_dimensions[2].height = 22
            for i, s in enumerate(sessions):
                tm = s.get("time_metrics",{})
                apply_row(ws, 3+i, [fmt_date(s.get("session_start","")), s.get("total_duration_formatted", s.get("session_duration_formatted","")), tm.get("active_formatted",""), tm.get("idle_formatted",""), tm.get("typing_efficiency_percent",0), s.get("break_count",0)], alt=(i%2==1))
            apply_row(ws, 3+n, ["AVERAGE","","","", round(sum(s.get("time_metrics",{}).get("typing_efficiency_percent",0) for s in sessions)/n,2), sum(s.get("break_count",0) for s in sessions)], total=True)
            set_widths(ws, [14,18,14,14,14,13])
            ws.freeze_panes = "A3"
            self.progress.emit(75)

            # ── Sheet 5: Breaks ────────────────────────────────────────────────
            self.log.emit("📝  Writing Breaks sheet...")
            ws = wb.create_sheet("Breaks")
            ws.sheet_view.showGridLines = False
            title_row(ws, 6, "Break Log — All Sessions")
            apply_header(ws, 2, ["Session Date","Break Start","Break End","Duration","Reason","Duration (seconds)"])
            ws.row_dimensions[2].height = 22
            row = 3; bi = 0
            for s in sessions:
                for b in s.get("breaks",[]):
                    apply_row(ws, row, [fmt_date(s.get("session_start","")), fmt_dt(b.get("start_time","")), fmt_dt(b.get("end_time","")), b.get("duration_formatted",""), b.get("reason",""), b.get("duration_seconds",0)], alt=(bi%2==1))
                    row += 1; bi += 1
            if bi > 0:
                total_secs = sum(b.get("duration_seconds",0) for s in sessions for b in s.get("breaks",[]))
                apply_row(ws, row, ["TOTAL","","","",f"{bi} breaks total", round(total_secs,2)], total=True)
            set_widths(ws, [14,20,20,14,22,18])
            ws.freeze_panes = "A3"
            self.progress.emit(88)

            # ── Sheet 6: Hourly ────────────────────────────────────────────────
            self.log.emit("📝  Writing Hourly Breakdown sheet...")
            ws = wb.create_sheet("Hourly Breakdown")
            ws.sheet_view.showGridLines = False
            title_row(ws, 10, "Hourly Keystroke Breakdown — All Sessions")
            apply_header(ws, 2, ["Session Date","Hour","Total Keys","Character Keys","Space Keys","Backspace Keys","Enter Keys","Modifier Keys","Special Keys","Active Seconds"])
            ws.row_dimensions[2].height = 22
            row = 3; hi = 0
            for s in sessions:
                for h in s.get("hourly_breakdown",[]):
                    apply_row(ws, row, [fmt_date(s.get("session_start","")), h.get("hour",""), h.get("total_keys",0), h.get("character_keys",0), h.get("space_keys",0), h.get("backspace_keys",0), h.get("enter_keys",0), h.get("modifier_keys",0), h.get("special_keys",0), h.get("active_seconds",0)], alt=(hi%2==1))
                    row += 1; hi += 1
            set_widths(ws, [14,8,12,14,12,15,12,14,13,15])
            ws.freeze_panes = "A3"
            self.progress.emit(96)

            # ── Save ───────────────────────────────────────────────────────────
            self.log.emit(f"\n💾  Saving to: {self.output_path}")
            Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(self.output_path)
            self.progress.emit(100)
            self.log.emit(f"✅  Done! {n} session(s) merged into {len(wb.sheetnames)} sheets.")
            self.finished.emit(self.output_path)

        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")


# ── Stat card widget ──────────────────────────────────────────────────────────
class StatCard(QFrame):
    def __init__(self, value, label):
        super().__init__()
        self.setObjectName("stat_card")
        self.setMinimumWidth(120)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(2)

        v = QLabel(str(value))
        v.setObjectName("stat_value")
        v.setAlignment(Qt.AlignCenter)

        l = QLabel(label)
        l.setObjectName("stat_label")
        l.setAlignment(Qt.AlignCenter)

        layout.addWidget(v)
        layout.addWidget(l)
        self.value_label = v

    def set_value(self, val):
        self.value_label.setText(str(val))


# ── Main window ───────────────────────────────────────────────────────────────
class AggregatorWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Session Aggregator — Keyboard Activity Tracker")
        self.setMinimumSize(780, 680)
        self.resize(860, 740)
        self._worker  = None
        self._thread  = None
        self._output_path = None
        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(16)

        # ── Header ────────────────────────────────────────────────────────────
        header = QFrame()
        header.setObjectName("card")
        hl = QVBoxLayout(header)
        hl.setContentsMargins(20, 16, 20, 16)
        hl.setSpacing(4)

        title = QLabel("⌨  Session Aggregator")
        title.setFont(QFont("Segoe UI", 18, QFont.Bold))
        title.setStyleSheet(f"color: {ACCENT}; background: transparent; border: none;")

        subtitle = QLabel("Merge Keyboard Activity Tracker JSON sessions into a formatted Excel report")
        subtitle.setStyleSheet(f"color: {TEXT_SECONDARY}; background: transparent; border: none; font-size: 12px;")

        hl.addWidget(title)
        hl.addWidget(subtitle)
        root.addWidget(header)

        # ── Input card ────────────────────────────────────────────────────────
        input_card = QFrame()
        input_card.setObjectName("card")
        il = QVBoxLayout(input_card)
        il.setContentsMargins(20, 16, 20, 16)
        il.setSpacing(12)

        sec1 = QLabel("CONFIGURATION")
        sec1.setObjectName("section_label")
        il.addWidget(sec1)

        # User ID
        uid_row = QHBoxLayout()
        uid_lbl = QLabel("Employee ID")
        uid_lbl.setFixedWidth(120)
        uid_lbl.setStyleSheet(f"color: {TEXT_SECONDARY}; background: transparent; border: none;")
        self.uid_edit = QLineEdit()
        self.uid_edit.setPlaceholderText("e.g. 7544")
        self.uid_edit.textChanged.connect(self._on_input_changed)
        uid_row.addWidget(uid_lbl)
        uid_row.addWidget(self.uid_edit)
        il.addLayout(uid_row)

        # Input folder
        in_row = QHBoxLayout()
        in_lbl = QLabel("Sessions Folder")
        in_lbl.setFixedWidth(120)
        in_lbl.setStyleSheet(f"color: {TEXT_SECONDARY}; background: transparent; border: none;")
        self.in_edit = QLineEdit()
        self.in_edit.setPlaceholderText(r"\\server\share\Keyboard_Activity_Tracker\7544  or  C:\sessions\7544")
        self.in_edit.textChanged.connect(self._on_input_changed)
        in_browse = QPushButton("Browse")
        in_browse.setObjectName("browse_btn")
        in_browse.setFixedWidth(90)
        in_browse.clicked.connect(self._browse_input)
        in_row.addWidget(in_lbl)
        in_row.addWidget(self.in_edit)
        in_row.addWidget(in_browse)
        il.addLayout(in_row)

        # Output file
        out_row = QHBoxLayout()
        out_lbl = QLabel("Output Excel")
        out_lbl.setFixedWidth(120)
        out_lbl.setStyleSheet(f"color: {TEXT_SECONDARY}; background: transparent; border: none;")
        self.out_edit = QLineEdit()
        self.out_edit.setPlaceholderText("e.g. C:\\Reports\\7544_report.xlsx")
        self.out_edit.textChanged.connect(self._on_input_changed)
        out_browse = QPushButton("Browse")
        out_browse.setObjectName("browse_btn")
        out_browse.setFixedWidth(90)
        out_browse.clicked.connect(self._browse_output)
        out_row.addWidget(out_lbl)
        out_row.addWidget(self.out_edit)
        out_row.addWidget(out_browse)
        il.addLayout(out_row)

        root.addWidget(input_card)

        # ── Preview stats ─────────────────────────────────────────────────────
        self.preview_card = QFrame()
        self.preview_card.setObjectName("card")
        pl = QVBoxLayout(self.preview_card)
        pl.setContentsMargins(20, 14, 20, 14)
        pl.setSpacing(10)

        sec2 = QLabel("SESSION PREVIEW")
        sec2.setObjectName("section_label")
        pl.addWidget(sec2)

        stats_row = QHBoxLayout()
        stats_row.setSpacing(10)
        self.stat_sessions  = StatCard("—", "Sessions Found")
        self.stat_keys      = StatCard("—", "Total Keys")
        self.stat_days      = StatCard("—", "Days Covered")
        self.stat_breaks    = StatCard("—", "Total Breaks")
        for sc in [self.stat_sessions, self.stat_keys, self.stat_days, self.stat_breaks]:
            stats_row.addWidget(sc)
        pl.addLayout(stats_row)
        self.preview_card.setVisible(False)
        root.addWidget(self.preview_card)

        # ── Progress + actions ────────────────────────────────────────────────
        action_card = QFrame()
        action_card.setObjectName("card")
        al = QVBoxLayout(action_card)
        al.setContentsMargins(20, 14, 20, 14)
        al.setSpacing(10)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setVisible(False)
        al.addWidget(self.progress_bar)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self.gen_btn = QPushButton("⚡  Generate Report")
        self.gen_btn.setObjectName("generate_btn")
        self.gen_btn.setEnabled(False)
        self.gen_btn.clicked.connect(self._generate)

        self.open_btn = QPushButton("📂  Open Excel")
        self.open_btn.setObjectName("open_btn")
        self.open_btn.setEnabled(False)
        self.open_btn.clicked.connect(self._open_output)

        btn_row.addWidget(self.gen_btn)
        btn_row.addWidget(self.open_btn)
        al.addLayout(btn_row)
        root.addWidget(action_card)

        # ── Log ───────────────────────────────────────────────────────────────
        log_card = QFrame()
        log_card.setObjectName("card")
        ll = QVBoxLayout(log_card)
        ll.setContentsMargins(20, 14, 20, 14)
        ll.setSpacing(8)

        sec3 = QLabel("LOG")
        sec3.setObjectName("section_label")
        ll.addWidget(sec3)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMinimumHeight(160)
        ll.addWidget(self.log_box)
        root.addWidget(log_card)

        self._log("Ready. Select a sessions folder and employee ID to begin.")

    # ── Browse handlers ───────────────────────────────────────────────────────
    def _browse_input(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Sessions Folder")
        if folder:
            self.in_edit.setText(folder)
            # Auto-fill employee ID from folder name
            name = Path(folder).name
            if name.isdigit() and not self.uid_edit.text():
                self.uid_edit.setText(name)
            # Auto-fill output path
            if not self.out_edit.text():
                uid = self.uid_edit.text() or name
                self.out_edit.setText(str(Path(folder).parent / f"{uid}_report.xlsx"))
            self._scan_preview()

    def _browse_output(self):
        uid = self.uid_edit.text() or "report"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report As",
            f"{uid}_report.xlsx",
            "Excel Files (*.xlsx)"
        )
        if path:
            self.out_edit.setText(path)

    # ── Preview scan ─────────────────────────────────────────────────────────
    def _scan_preview(self):
        folder = self.in_edit.text().strip()
        if not folder or not Path(folder).exists():
            self.preview_card.setVisible(False)
            return
        try:
            files = list(Path(folder).glob("*.json"))
            sessions = []
            for f in files:
                try:
                    d = json.loads(f.read_text(encoding="utf-8"))
                    if not d.get("encrypted"):
                        sessions.append(d)
                except Exception:
                    pass
            if sessions:
                total_keys = sum(s.get("keystroke_counts",{}).get("total",0) for s in sessions)
                days = len(set(s.get("session_start","")[:10] for s in sessions))
                total_breaks = sum(s.get("break_count",0) for s in sessions)
                self.stat_sessions.set_value(len(sessions))
                self.stat_keys.set_value(f"{total_keys:,}")
                self.stat_days.set_value(days)
                self.stat_breaks.set_value(total_breaks)
                self.preview_card.setVisible(True)
            else:
                self.preview_card.setVisible(False)
        except Exception:
            self.preview_card.setVisible(False)

    # ── Input validation ──────────────────────────────────────────────────────
    def _on_input_changed(self):
        uid  = self.uid_edit.text().strip()
        inp  = self.in_edit.text().strip()
        out  = self.out_edit.text().strip()
        self.gen_btn.setEnabled(bool(uid and inp and out))
        if inp:
            self._scan_preview()

    # ── Generate ──────────────────────────────────────────────────────────────
    def _generate(self):
        uid = self.uid_edit.text().strip()
        inp = self.in_edit.text().strip()
        out = self.out_edit.text().strip()

        if not out.endswith(".xlsx"):
            out += ".xlsx"
            self.out_edit.setText(out)

        self.gen_btn.setEnabled(False)
        self.open_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.log_box.clear()
        self._log(f"Starting report generation for user: {uid}")

        self._worker = ReportWorker(inp, out, uid)
        self._thread = QThread()
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._on_progress)
        self._worker.log.connect(self._log)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.finished.connect(self._thread.quit)
        self._worker.error.connect(self._thread.quit)
        self._thread.start()

    # ── Worker signals ────────────────────────────────────────────────────────
    def _on_progress(self, val):
        self.progress_bar.setValue(val)

    def _on_finished(self, path):
        self._output_path = path
        self.open_btn.setEnabled(True)
        self.gen_btn.setEnabled(True)
        self._log(f"\n🎉  Report ready: {path}")
        self.progress_bar.setStyleSheet(
            f"QProgressBar::chunk {{ background-color: {ACCENT_GREEN}; border-radius: 4px; }}"
        )

    def _on_error(self, msg):
        self._log(f"\n❌  ERROR: {msg}")
        self.gen_btn.setEnabled(True)
        self.progress_bar.setStyleSheet(
            f"QProgressBar::chunk {{ background-color: {ACCENT_RED}; border-radius: 4px; }}"
        )

    def _open_output(self):
        if self._output_path and Path(self._output_path).exists():
            os.startfile(self._output_path)

    def _log(self, msg):
        self.log_box.append(msg)
        self.log_box.verticalScrollBar().setValue(
            self.log_box.verticalScrollBar().maximum()
        )


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)
    app.setStyle("Fusion")
    window = AggregatorWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()