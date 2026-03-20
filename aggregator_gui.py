#!/usr/bin/env python3
"""
Keyboard Productivity Tracking Tool - Session-by-Session Report Generator v2.0
Integrates with Employee_Project_Master for business context enrichment.
Supports enhanced JSON format v2.0 with metadata, break_reasons, and idle_periods.
"""
import json
import os
import re
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")
    print("Install with: pip install openpyxl")

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QTextEdit, QFileDialog,
    QProgressBar, QGroupBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QTabWidget, QMessageBox, QSplitter, QComboBox, QCheckBox
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
COLOR_HEADER_BLUE   = "1F4E78"
COLOR_HEADER_GREEN  = "4CAF50"
COLOR_HEADER_ORANGE = "FF9800"
COLOR_HEADER_RED    = "F44336"
COLOR_HEADER_PURPLE = "6A0DAD"
COLOR_HEADER_TEAL   = "008080"
COLOR_UNKNOWN_BG    = "FFF4CE"
COLOR_WHITE_FONT    = "FFFFFF"

THIN_BORDER = None   # built lazily in _thin_border()


def _thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(cell, fill_color, font_size=11):
    cell.fill   = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font   = Font(bold=True, color=COLOR_WHITE_FONT, size=font_size, name="Arial")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _thin_border()


def _data_cell(cell, value, highlight_unknown=False):
    cell.value = value
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = Font(name="Arial", size=10)
    if highlight_unknown and value == 'UNKNOWN':
        cell.fill = PatternFill(start_color=COLOR_UNKNOWN_BG, end_color=COLOR_UNKNOWN_BG,
                                fill_type="solid")


# ---------------------------------------------------------------------------
# Employee Master Loader
# ---------------------------------------------------------------------------
class EmployeeMasterLoader:
    REQUIRED_COLUMNS = [
        'PSN', 'Employee_Name', 'Team_Lead', 'Project_Code',
        'Project_Type', 'Role', 'Shift', 'Location'
    ]

    def __init__(self, master_file_path: str):
        self.master_file_path = Path(master_file_path)
        self.employee_data: Dict[str, Dict] = {}

    def load(self) -> Dict[str, Dict]:
        if not self.master_file_path.exists():
            raise FileNotFoundError(f"Master file not found: {self.master_file_path}")
        df = pd.read_excel(self.master_file_path)
        missing = [c for c in self.REQUIRED_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"Missing columns in master file: {missing}")
        for _, row in df.iterrows():
            psn = str(row['PSN']).strip()
            self.employee_data[psn] = {k: str(row[k]).strip() for k in self.REQUIRED_COLUMNS[1:]}
        return self.employee_data

    def get_employee_info(self, psn: str) -> Dict:
        return self.employee_data.get(psn, {k: 'UNKNOWN' for k in self.REQUIRED_COLUMNS[1:]}).copy()


# ---------------------------------------------------------------------------
# Worker Thread
# ---------------------------------------------------------------------------
class AggregatorWorker(QThread):
    progress_update = pyqtSignal(int, str)
    finished        = pyqtSignal(dict)
    error           = pyqtSignal(str)

    def __init__(self, source_path: str, output_path: str, master_file_path: str = None):
        super().__init__()
        self.source_path       = Path(source_path)
        self.output_path       = Path(output_path)
        self.master_file_path  = master_file_path
        self.filename_pattern  = re.compile(r'^(\w+)_(\d{8})_(\d{6})\.json$')
        self.failed_files: List[Dict] = []
        self.employee_master: Optional[EmployeeMasterLoader] = None

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def seconds_to_hhmm(seconds: float) -> str:
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        return f"{h:02d}:{m:02d}"

    @staticmethod
    def seconds_to_hhmmss(seconds: float) -> str:
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        s = int(seconds % 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    @staticmethod
    def format_time(time_str: str) -> str:
        try:
            return datetime.fromisoformat(time_str).strftime('%H:%M:%S')
        except Exception:
            return time_str or '--:--:--'

    # ------------------------------------------------------------------
    # Thread entry
    # ------------------------------------------------------------------
    def run(self):
        try:
            if self.master_file_path:
                self.progress_update.emit(0, "Loading Employee Master data...")
                self.employee_master = EmployeeMasterLoader(self.master_file_path)
                self.employee_master.load()
                self.progress_update.emit(5, f"Loaded {len(self.employee_master.employee_data)} employee records")

            sessions = self.process_files()
            summary  = self.create_summary(sessions)
            self.save_to_excel(summary)
            self.finished.emit(summary)
        except Exception as e:
            self.error.emit(str(e))

    # ------------------------------------------------------------------
    # File processing
    # ------------------------------------------------------------------
    def parse_filename(self, filename: str) -> Optional[Tuple[str, str, str]]:
        m = self.filename_pattern.match(filename)
        return m.groups() if m else None

    def read_json_file(self, filepath: Path) -> Optional[Dict]:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            self.failed_files.append({'file': filepath.name, 'error': f'JSON decode error: {e}'})
        except Exception as e:
            self.failed_files.append({'file': filepath.name, 'error': f'Read error: {e}'})
        return None

    def extract_session_data(self, data: Dict, filename: str) -> Optional[Dict]:
        try:
            session  = data.get('session', {})
            metadata = data.get('metadata', {})

            # Prefer username from metadata; fall back to filename stem
            psn_meta = str(metadata.get('username', '')).strip()

            return {
                'filename':             filename,
                # --- metadata fields ---
                'psn_from_metadata':    psn_meta,
                'computer_name':        metadata.get('computer_name', ''),
                'version':              metadata.get('version', ''),
                'storage_location':     metadata.get('storage_location', ''),
                'always_on_mode':       metadata.get('always_on_mode', False),
                # --- session identity ---
                'session_id':           session.get('session_id', ''),
                'start_time':           session.get('start_time', ''),
                'end_time':             session.get('end_time', ''),
                # --- durations ---
                'duration':             session.get('total_duration', 0),
                'active_time':          session.get('active_time', 0),
                'idle_time':            session.get('idle_time', 0),
                'current_idle_duration':session.get('current_idle_duration', 0),
                # --- keystroke metrics ---
                'total_keys':           session.get('total_keys', 0),
                'average_kpm':          session.get('average_kpm', 0),
                'active_kpm':           session.get('active_kpm', 0),
                'peak_kpm':             session.get('peak_kpm', 0),
                'typing_efficiency':    session.get('typing_efficiency', 0),
                # --- idle stats ---
                'idle_periods_count':   session.get('idle_periods_count', 0),
                'longest_idle_period':  session.get('longest_idle_period', 0),
                'average_idle_period':  session.get('average_idle_period', 0),
                # --- structured data ---
                'key_categories':       session.get('key_categories', {}),
                'break_reasons':        session.get('break_reasons', {}),
                'idle_periods':         session.get('idle_periods', []),
            }
        except Exception:
            return None

    def process_files(self) -> List[Dict]:
        sessions: List[Dict] = []
        json_files = list(self.source_path.glob('*.json'))
        total = len(json_files)
        self.progress_update.emit(5, f"Found {total} JSON files")

        for idx, filepath in enumerate(json_files, 1):
            parsed = self.parse_filename(filepath.name)
            if not parsed:
                self.failed_files.append({'file': filepath.name, 'error': 'Invalid filename format'})
                continue

            psn_filename, date_str, time_str = parsed
            data = self.read_json_file(filepath)
            if not data:
                continue

            metrics = self.extract_session_data(data, filepath.name)
            if not metrics:
                self.failed_files.append({'file': filepath.name, 'error': 'Invalid session data structure'})
                continue

            # Use metadata PSN if available, else fall back to filename PSN
            effective_psn = metrics['psn_from_metadata'] or psn_filename

            formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
            session_time   = f"{time_str[:2]}:{time_str[2:4]}:{time_str[4:]}"

            sessions.append({
                'psn':          effective_psn,
                'psn_filename': psn_filename,
                'date':         formatted_date,
                'session_time': session_time,
                **metrics
            })

            self.progress_update.emit(int(5 + (idx / total) * 85), f"Processing: {filepath.name}")

        return sessions

    # ------------------------------------------------------------------
    # Summary creation
    # ------------------------------------------------------------------
    def create_summary(self, sessions: List[Dict]) -> Dict:
        unknown_info = {k: 'UNKNOWN' for k in [
            'Employee_Name', 'Team_Lead', 'Project_Code',
            'Project_Type', 'Role', 'Shift', 'Location'
        ]}

        enriched = []
        for s in sessions:
            emp = self.employee_master.get_employee_info(s['psn']) if self.employee_master else unknown_info.copy()
            enriched.append({**s, 'employee_info': emp})

        enriched.sort(key=lambda x: (x['psn'], x['date'], x['session_time']))

        summary = {
            'generated_at': datetime.now().isoformat(),
            'statistics': {
                'successfully_processed': len(enriched),
                'failed_files':           len(self.failed_files),
                'total_users':            len(set(s['psn'] for s in enriched)),
                'total_unique_dates':     len(set(s['date'] for s in enriched)),
                'master_file_loaded':     bool(self.employee_master),
            },
            'sessions': enriched,
        }

        if self.failed_files:
            summary['failed_files'] = self.failed_files

        return summary

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def save_to_excel(self, summary: Dict):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self.create_productivity_report_sheet(wb, summary)
        self.create_summary_sheet(wb, summary)
        self.create_key_categories_sheet(wb, summary)
        self.create_break_reasons_sheet(wb, summary)
        self.create_idle_periods_sheet(wb, summary)

        if summary.get('failed_files'):
            self.create_failed_files_sheet(wb, summary)

        wb.save(self.output_path)

    # ------ Sheet 1: Productivity Report ----------------------------------
    def create_productivity_report_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Productivity Report", 0)

        headers = [
            'PSN', 'Employee_Name', 'Team_Lead', 'Project_Code', 'Project_Type',
            'Role', 'Shift', 'Location',
            'Date', 'Session_Time', 'Session_ID',
            'Start_Time', 'End_Time',
            'Duration (HH:MM)', 'Active_Time (HH:MM)', 'Idle_Time (HH:MM)',
            'Total_Keys', 'Avg_KPM', 'Active_KPM', 'Peak_KPM',
            'Typing_Efficiency (%)',
            'Idle_Count', 'Longest_Idle (HH:MM:SS)', 'Avg_Idle (HH:MM:SS)',
            'Computer_Name', 'Storage_Location', 'Always_On_Mode',
            'Filename'
        ]

        for col, h in enumerate(headers, 1):
            _header_cell(ws.cell(row=1, column=col), COLOR_HEADER_BLUE)
            ws.cell(row=1, column=col).value = h

        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 30

        for row_num, s in enumerate(summary['sessions'], 2):
            emp = s.get('employee_info', {})
            vals = [
                s['psn'],
                emp.get('Employee_Name', 'UNKNOWN'),
                emp.get('Team_Lead', 'UNKNOWN'),
                emp.get('Project_Code', 'UNKNOWN'),
                emp.get('Project_Type', 'UNKNOWN'),
                emp.get('Role', 'UNKNOWN'),
                emp.get('Shift', 'UNKNOWN'),
                emp.get('Location', 'UNKNOWN'),
                s['date'],
                s['session_time'],
                s.get('session_id', ''),
                self.format_time(s.get('start_time', '')),
                self.format_time(s.get('end_time', '')),
                self.seconds_to_hhmm(s['duration']),
                self.seconds_to_hhmm(s['active_time']),
                self.seconds_to_hhmm(s['idle_time']),
                s['total_keys'],
                round(s['average_kpm'], 2),
                round(s['active_kpm'], 2),
                round(s['peak_kpm'], 2),
                round(s['typing_efficiency'], 2),
                s['idle_periods_count'],
                self.seconds_to_hhmmss(s['longest_idle_period']),
                self.seconds_to_hhmmss(s['average_idle_period']),
                s.get('computer_name', ''),
                s.get('storage_location', ''),
                'Yes' if s.get('always_on_mode') else 'No',
                s['filename'],
            ]
            for col, v in enumerate(vals, 1):
                _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))

        widths = [12,20,18,15,15,12,10,15,12,13,18,12,12,16,16,16,12,10,10,10,18,10,18,18,20,15,14,30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------ Sheet 2: Summary ----------------------------------------------
    def create_summary_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Summary")
        stats = summary['statistics']

        ws['A1'] = "Productivity Aggregation Summary"
        ws['A1'].font = Font(bold=True, size=16, name="Arial")
        ws.merge_cells('A1:C1')

        meta_rows = [
            ("Generated At",              summary['generated_at']),
            ("Employee Master Loaded",     "Yes" if stats['master_file_loaded'] else "No"),
            ("Total Users",                stats['total_users']),
            ("Total Unique Dates",         stats['total_unique_dates']),
            ("Sessions Processed",         stats['successfully_processed']),
            ("Failed Files",               stats['failed_files']),
        ]
        for r, (k, v) in enumerate(meta_rows, 3):
            ws.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial")
            ws.cell(row=r, column=2, value=v).font = Font(name="Arial")

        start_row = 3 + len(meta_rows) + 2
        headers = ['PSN','Employee_Name','Team_Lead','Project_Code','Project_Type',
                   'Role','Shift','Location','Total_Sessions','Total_Keys',
                   'Total_Active_Time (HH:MM)','Avg_Typing_Efficiency (%)']

        for col, h in enumerate(headers, 1):
            _header_cell(ws.cell(row=start_row, column=col), COLOR_HEADER_GREEN)
            ws.cell(row=start_row, column=col).value = h

        agg = defaultdict(lambda: {'sessions':0,'total_keys':0,'active_time':0.0,
                                   'efficiency_sum':0.0,'employee_info':None})
        for s in summary['sessions']:
            p = s['psn']
            agg[p]['sessions']      += 1
            agg[p]['total_keys']    += s['total_keys']
            agg[p]['active_time']   += s['active_time']
            agg[p]['efficiency_sum']+= s['typing_efficiency']
            if agg[p]['employee_info'] is None:
                agg[p]['employee_info'] = s['employee_info']

        for row_num, (psn, d) in enumerate(sorted(agg.items()), start_row + 1):
            emp = d['employee_info'] or {}
            avg_eff = d['efficiency_sum'] / d['sessions'] if d['sessions'] else 0
            vals = [
                psn,
                emp.get('Employee_Name','UNKNOWN'), emp.get('Team_Lead','UNKNOWN'),
                emp.get('Project_Code','UNKNOWN'),  emp.get('Project_Type','UNKNOWN'),
                emp.get('Role','UNKNOWN'),           emp.get('Shift','UNKNOWN'),
                emp.get('Location','UNKNOWN'),
                d['sessions'], d['total_keys'],
                self.seconds_to_hhmm(d['active_time']),
                round(avg_eff, 2),
            ]
            for col, v in enumerate(vals, 1):
                _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))

        widths = [12,20,18,15,15,12,10,15,14,14,22,22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------ Sheet 3: Key Categories ---------------------------------------
    def create_key_categories_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Key Categories")

        all_cats = sorted({cat for s in summary['sessions'] for cat in s['key_categories']})

        base_headers = ['PSN','Employee_Name','Team_Lead','Project_Code','Project_Type',
                        'Role','Shift','Location','Date','Session_Time','Session_ID']
        headers = base_headers + all_cats

        for col, h in enumerate(headers, 1):
            _header_cell(ws.cell(row=1, column=col), COLOR_HEADER_ORANGE)
            ws.cell(row=1, column=col).value = h

        ws.freeze_panes = 'A2'

        for row_num, s in enumerate(summary['sessions'], 2):
            emp = s.get('employee_info', {})
            base_vals = [
                s['psn'], emp.get('Employee_Name','UNKNOWN'), emp.get('Team_Lead','UNKNOWN'),
                emp.get('Project_Code','UNKNOWN'), emp.get('Project_Type','UNKNOWN'),
                emp.get('Role','UNKNOWN'), emp.get('Shift','UNKNOWN'), emp.get('Location','UNKNOWN'),
                s['date'], s['session_time'], s.get('session_id',''),
            ]
            cat_vals = [s['key_categories'].get(c, 0) for c in all_cats]
            for col, v in enumerate(base_vals + cat_vals, 1):
                _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))

        base_widths = [12,20,18,15,15,12,10,15,12,13,18]
        for i, w in enumerate(base_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for i in range(len(base_headers)+1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    # ------ Sheet 4: Break Reasons ----------------------------------------
    def create_break_reasons_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Break Reasons")

        headers = [
            'PSN','Employee_Name','Team_Lead','Project_Code','Project_Type',
            'Role','Shift','Location','Date','Session_Time','Session_ID',
            'Break_Reason','Break_Count','Total_Duration (HH:MM:SS)','Total_Duration (Seconds)'
        ]

        for col, h in enumerate(headers, 1):
            _header_cell(ws.cell(row=1, column=col), COLOR_HEADER_PURPLE)
            ws.cell(row=1, column=col).value = h

        ws.freeze_panes = 'A2'

        row_num = 2
        for s in summary['sessions']:
            emp = s.get('employee_info', {})
            base = [
                s['psn'], emp.get('Employee_Name','UNKNOWN'), emp.get('Team_Lead','UNKNOWN'),
                emp.get('Project_Code','UNKNOWN'), emp.get('Project_Type','UNKNOWN'),
                emp.get('Role','UNKNOWN'), emp.get('Shift','UNKNOWN'), emp.get('Location','UNKNOWN'),
                s['date'], s['session_time'], s.get('session_id',''),
            ]
            br = s.get('break_reasons', {})
            if not br:
                # Still write one row with no break data
                vals = base + ['N/A', 0, '00:00:00', 0]
                for col, v in enumerate(vals, 1):
                    _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))
                row_num += 1
            else:
                for reason, rd in br.items():
                    cnt  = rd.get('count', 0)
                    dur  = rd.get('total_duration', 0)
                    vals = base + [reason, cnt, self.seconds_to_hhmmss(dur), round(dur, 2)]
                    for col, v in enumerate(vals, 1):
                        _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))
                    row_num += 1

        widths = [12,20,18,15,15,12,10,15,12,13,18,30,12,20,20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------ Sheet 5: Idle Periods -----------------------------------------
    def create_idle_periods_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Idle Periods")

        headers = [
            'PSN','Employee_Name','Team_Lead','Project_Code','Project_Type',
            'Role','Shift','Location','Date','Session_Time','Session_ID',
            'Idle_Start','Idle_End','Duration (HH:MM:SS)','Duration (Seconds)',
            'Break_Reason','Details'
        ]

        for col, h in enumerate(headers, 1):
            _header_cell(ws.cell(row=1, column=col), COLOR_HEADER_TEAL)
            ws.cell(row=1, column=col).value = h

        ws.freeze_panes = 'A2'

        row_num = 2
        for s in summary['sessions']:
            emp = s.get('employee_info', {})
            base = [
                s['psn'], emp.get('Employee_Name','UNKNOWN'), emp.get('Team_Lead','UNKNOWN'),
                emp.get('Project_Code','UNKNOWN'), emp.get('Project_Type','UNKNOWN'),
                emp.get('Role','UNKNOWN'), emp.get('Shift','UNKNOWN'), emp.get('Location','UNKNOWN'),
                s['date'], s['session_time'], s.get('session_id',''),
            ]
            idle_list = s.get('idle_periods', [])
            if not idle_list:
                vals = base + ['','','00:00:00', 0,'','']
                for col, v in enumerate(vals, 1):
                    _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))
                row_num += 1
            else:
                for ip in idle_list:
                    dur    = ip.get('duration', 0)
                    reason = ip.get('reason') or 'Not specified'
                    vals   = base + [
                        self.format_time(ip.get('start', '')),
                        self.format_time(ip.get('end', '')),
                        self.seconds_to_hhmmss(dur),
                        round(dur, 2),
                        reason,
                        ip.get('details', ''),
                    ]
                    for col, v in enumerate(vals, 1):
                        _data_cell(ws.cell(row=row_num, column=col), v, highlight_unknown=(col <= 8))
                    row_num += 1

        widths = [12,20,18,15,15,12,10,15,12,13,18,12,12,18,16,30,20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------ Sheet 6: Failed Files -----------------------------------------
    def create_failed_files_sheet(self, wb, summary: Dict):
        ws = wb.create_sheet("Failed Files")

        for col, h in enumerate(['File Name', 'Error'], 1):
            _header_cell(ws.cell(row=1, column=col), COLOR_HEADER_RED)
            ws.cell(row=1, column=col).value = h

        for row_num, f in enumerate(summary.get('failed_files', []), 2):
            ws.cell(row=row_num, column=1, value=f['file']).font = Font(name="Arial", size=10)
            ws.cell(row=row_num, column=2, value=f['error']).font = Font(name="Arial", size=10)

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 65


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------
class SessionAggregatorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.worker       = None
        self.summary_data = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("Keyboard Productivity Session Aggregator v2.0")
        self.setGeometry(100, 100, 1300, 900)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        layout.addWidget(self.create_path_selection())
        layout.addWidget(self.create_progress_section())

        # Results tabs
        self.results_tabs = QTabWidget()
        self.results_tabs.setEnabled(False)

        # Statistics
        self.stats_text = QTextEdit()
        self.stats_text.setReadOnly(True)
        self.stats_text.setFont(QFont("Courier New", 10))
        self.results_tabs.addTab(self.stats_text, "Statistics")

        # Productivity preview
        preview_headers = [
            "PSN","Employee","Team Lead","Project","Type","Role","Shift","Location",
            "Date","Session Time","Session ID","Start","End",
            "Duration","Active","Idle","Keys","Avg KPM","Active KPM","Peak KPM",
            "Efficiency %","Idle Count","Longest Idle","Avg Idle",
            "Computer","Storage","Always On","Filename"
        ]
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(len(preview_headers))
        self.report_table.setHorizontalHeaderLabels(preview_headers)
        self.report_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.results_tabs.addTab(self.report_table, "Productivity Report Preview")

        # Failed files
        self.failed_text = QTextEdit()
        self.failed_text.setReadOnly(True)
        self.failed_text.setFont(QFont("Courier New", 9))
        self.results_tabs.addTab(self.failed_text, "Failed Files")

        layout.addWidget(self.results_tabs)
        self.apply_stylesheet()

    def apply_stylesheet(self):
        self.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #cccccc;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                background-color: #1F4E78;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover    { background-color: #2E75B6; }
            QPushButton:disabled { background-color: #cccccc; color: #666666; }
            QLineEdit {
                padding: 5px;
                border: 1px solid #cccccc;
                border-radius: 3px;
            }
            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 3px;
                text-align: center;
            }
            QProgressBar::chunk { background-color: #1F4E78; }
        """)

    def create_path_selection(self) -> QGroupBox:
        group  = QGroupBox("File Selection")
        layout = QVBoxLayout()

        def make_row(label, placeholder, browse_slot):
            row   = QHBoxLayout()
            lbl   = QLabel(label)
            lbl.setFixedWidth(160)
            edit  = QLineEdit()
            edit.setPlaceholderText(placeholder)
            btn   = QPushButton("Browse")
            btn.setFixedWidth(80)
            btn.clicked.connect(browse_slot)
            row.addWidget(lbl)
            row.addWidget(edit)
            row.addWidget(btn)
            return row, edit, btn

        r1, self.master_path_edit, self.browse_master_btn = make_row(
            "Employee Master:", "Select Employee_Project_Master.xlsx...", self.browse_master)
        r2, self.source_path_edit, self.browse_source_btn = make_row(
            "Session Files Folder:", "Select folder containing JSON session files...", self.browse_source)
        r3, self.output_path_edit, self.browse_output_btn = make_row(
            "Output Excel File:", "Productivity_Report_v2.xlsx", self.browse_output)
        self.output_path_edit.setText("Productivity_Report_v2.xlsx")

        layout.addLayout(r1)
        layout.addLayout(r2)
        layout.addLayout(r3)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.start_btn = QPushButton("Generate Productivity Report")
        self.start_btn.setMinimumWidth(260)
        self.start_btn.clicked.connect(self.start_aggregation)
        btn_row.addWidget(self.start_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        group.setLayout(layout)
        return group

    def create_progress_section(self) -> QGroupBox:
        group  = QGroupBox("Progress")
        layout = QVBoxLayout()
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.status_label = QLabel("Ready to process files")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.status_label)
        group.setLayout(layout)
        return group

    # ------------------------------------------------------------------
    # Browse slots
    # ------------------------------------------------------------------
    def browse_master(self):
        p, _ = QFileDialog.getOpenFileName(self, "Select Employee Master File", "",
                                           "Excel Files (*.xlsx *.xls);;All Files (*)")
        if p:
            self.master_path_edit.setText(p)

    def browse_source(self):
        f = QFileDialog.getExistingDirectory(self, "Select Session Files Folder", "",
                                             QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks)
        if f:
            self.source_path_edit.setText(f)

    def browse_output(self):
        p, _ = QFileDialog.getSaveFileName(self, "Save Productivity Report",
                                           "Productivity_Report_v2.xlsx",
                                           "Excel Files (*.xlsx);;All Files (*)")
        if p:
            self.output_path_edit.setText(p)

    # ------------------------------------------------------------------
    # Run
    # ------------------------------------------------------------------
    def set_controls_enabled(self, enabled: bool):
        for w in [self.start_btn, self.browse_master_btn,
                  self.browse_source_btn, self.browse_output_btn]:
            w.setEnabled(enabled)

    def start_aggregation(self):
        master = self.master_path_edit.text().strip()
        source = self.source_path_edit.text().strip()
        output = self.output_path_edit.text().strip()

        if not source:
            QMessageBox.warning(self, "Input Error", "Please select a session files folder."); return
        if not output:
            QMessageBox.warning(self, "Input Error", "Please specify an output file."); return
        if not Path(source).exists():
            QMessageBox.warning(self, "Path Error", "Source folder does not exist."); return
        if master and not Path(master).exists():
            QMessageBox.warning(self, "Path Error", "Employee master file does not exist."); return
        if not EXCEL_AVAILABLE:
            QMessageBox.warning(self, "Missing Libraries",
                                "Excel export requires openpyxl and pandas.\n"
                                "Install: pip install openpyxl pandas"); return

        if not master:
            reply = QMessageBox.question(
                self, "No Master File",
                "No Employee Master file selected.\nAll employee fields will show as UNKNOWN.\n\nContinue?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No:
                return

        self.set_controls_enabled(False)
        self.results_tabs.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting aggregation...")

        self.worker = AggregatorWorker(source, output, master or None)
        self.worker.progress_update.connect(self.update_progress)
        self.worker.finished.connect(self.aggregation_finished)
        self.worker.error.connect(self.aggregation_error)
        self.worker.start()

    def update_progress(self, value: int, message: str):
        self.progress_bar.setValue(value)
        self.status_label.setText(message)

    def aggregation_finished(self, summary: Dict):
        self.summary_data = summary
        self.progress_bar.setValue(100)
        self.status_label.setText("Report generation completed successfully!")
        self.set_controls_enabled(True)
        self.results_tabs.setEnabled(True)

        self.display_statistics(summary)
        self.display_report_preview(summary)
        self.display_failed_files(summary)

        unknown_count = sum(
            1 for s in summary['sessions']
            if s.get('employee_info', {}).get('Employee_Name') == 'UNKNOWN'
        )
        stats = summary['statistics']
        QMessageBox.information(self, "Success",
            f"Productivity Report generated!\n\n"
            f"Sessions processed : {stats['successfully_processed']}\n"
            f"Unique users       : {stats['total_users']}\n"
            f"UNKNOWN sessions   : {unknown_count}\n"
            f"Failed files       : {stats['failed_files']}\n\n"
            f"Saved to:\n{self.output_path_edit.text()}"
        )

    def aggregation_error(self, error_msg: str):
        self.status_label.setText("Error during aggregation.")
        self.set_controls_enabled(True)
        QMessageBox.critical(self, "Error", f"Report generation failed:\n\n{error_msg}")

    # ------------------------------------------------------------------
    # Results display
    # ------------------------------------------------------------------
    def display_statistics(self, summary: Dict):
        stats = summary['statistics']
        lines = [
            "=" * 70,
            "PRODUCTIVITY REPORT STATISTICS (v2.0)",
            "=" * 70,
            f"Generated At          : {summary['generated_at']}",
            f"Employee Master Loaded: {'Yes' if stats['master_file_loaded'] else 'No'}",
            f"Sessions Processed    : {stats['successfully_processed']}",
            f"Unique Users          : {stats['total_users']}",
            f"Unique Dates          : {stats['total_unique_dates']}",
            f"Failed Files          : {stats['failed_files']}",
            "",
            "=" * 70,
            "SAMPLE SESSIONS (first 10)",
            "=" * 70,
        ]
        for i, s in enumerate(summary['sessions'][:10], 1):
            emp  = s.get('employee_info', {})
            name = emp.get('Employee_Name', 'UNKNOWN')
            eff  = s.get('typing_efficiency', 0)
            lines += [
                f"{i:2}. PSN: {s['psn']} | {name}",
                f"    Date: {s['date']}  Session: {s['session_time']}  ID: {s.get('session_id','')}",
                f"    Keys: {s['total_keys']:,}  Active KPM: {s.get('active_kpm',0):.1f}"
                f"  Efficiency: {eff:.1f}%",
                f"    Idle count: {s.get('idle_periods_count',0)}"
                f"  Break reasons: {list(s.get('break_reasons',{}).keys())}",
                "",
            ]
        self.stats_text.setText("\n".join(lines))

    @staticmethod
    def seconds_to_hhmm(seconds: float) -> str:
        h = int(seconds // 3600); m = int((seconds % 3600) // 60)
        return f"{h:02d}:{m:02d}"

    @staticmethod
    def seconds_to_hhmmss(seconds: float) -> str:
        h = int(seconds // 3600); m = int((seconds % 3600) // 60); s = int(seconds % 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    @staticmethod
    def format_time(time_str: str) -> str:
        try:
            return datetime.fromisoformat(time_str).strftime('%H:%M:%S')
        except Exception:
            return time_str or '--:--:--'

    def display_report_preview(self, summary: Dict):
        self.report_table.setRowCount(0)
        MAX = 100
        for row, s in enumerate(summary['sessions'][:MAX]):
            emp = s.get('employee_info', {})
            self.report_table.insertRow(row)
            vals = [
                s['psn'],
                emp.get('Employee_Name','UNKNOWN'), emp.get('Team_Lead','UNKNOWN'),
                emp.get('Project_Code','UNKNOWN'),  emp.get('Project_Type','UNKNOWN'),
                emp.get('Role','UNKNOWN'),           emp.get('Shift','UNKNOWN'),
                emp.get('Location','UNKNOWN'),
                s['date'], s['session_time'], s.get('session_id',''),
                self.format_time(s.get('start_time','')),
                self.format_time(s.get('end_time','')),
                self.seconds_to_hhmm(s['duration']),
                self.seconds_to_hhmm(s['active_time']),
                self.seconds_to_hhmm(s['idle_time']),
                str(s['total_keys']),
                f"{s['average_kpm']:.2f}",
                f"{s.get('active_kpm',0):.2f}",
                f"{s['peak_kpm']:.2f}",
                f"{s.get('typing_efficiency',0):.2f}",
                str(s.get('idle_periods_count',0)),
                self.seconds_to_hhmmss(s.get('longest_idle_period',0)),
                self.seconds_to_hhmmss(s.get('average_idle_period',0)),
                s.get('computer_name',''),
                s.get('storage_location',''),
                'Yes' if s.get('always_on_mode') else 'No',
                s['filename'],
            ]
            for col, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                if v == 'UNKNOWN':
                    item.setBackground(QColor(255, 244, 206))
                self.report_table.setItem(row, col, item)

        if len(summary['sessions']) > MAX:
            QMessageBox.information(self, "Preview Limit",
                f"Showing first {MAX} sessions. Full data is in the Excel file.\n"
                f"Total sessions: {len(summary['sessions'])}")
        self.report_table.resizeColumnsToContents()

    def display_failed_files(self, summary: Dict):
        ff = summary.get('failed_files', [])
        if ff:
            lines = ["FAILED FILES", "=" * 70, ""]
            for f in ff:
                lines += [f"File : {f['file']}", f"Error: {f['error']}", ""]
            self.failed_text.setText("\n".join(lines))
        else:
            self.failed_text.setText("No failed files — all sessions processed successfully!")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def main():
    import sys
    missing = []
    if not EXCEL_AVAILABLE:
        missing.append("openpyxl")
    try:
        import pandas
    except ImportError:
        missing.append("pandas")

    if missing:
        print(f"ERROR: Missing dependencies: {', '.join(missing)}")
        print(f"Run: pip install {' '.join(missing)}")
        sys.exit(1)

    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    gui = SessionAggregatorGUI()
    gui.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()